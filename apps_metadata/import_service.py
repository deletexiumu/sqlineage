import json
import csv
import pandas as pd
from io import StringIO
from typing import List, Dict, Any, Tuple
from django.db import transaction
from .models import HiveTable


class MetadataImportService:
    """元数据导入服务类"""
    
    SUPPORTED_FORMATS = ['json', 'csv', 'excel']
    
    def __init__(self):
        self.validation_errors = []
        self.import_stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'errors': 0,
            'skipped': 0
        }
    
    def validate_table_data(self, table_data: Dict[str, Any]) -> bool:
        """验证单个表的数据格式"""
        errors = []
        
        # 必需字段检查
        required_fields = ['database', 'name', 'columns']
        for field in required_fields:
            if field not in table_data:
                errors.append(f"缺少必需字段: {field}")
        
        # 数据库名和表名长度检查
        if 'database' in table_data and len(table_data['database']) > 255:
            errors.append("数据库名称过长（最大255字符）")
        
        if 'name' in table_data and len(table_data['name']) > 255:
            errors.append("表名过长（最大255字符）")
        
        # 列数据验证
        if 'columns' in table_data:
            if not isinstance(table_data['columns'], list):
                errors.append("columns 字段必须是数组格式")
            else:
                for i, column in enumerate(table_data['columns']):
                    if not isinstance(column, dict):
                        errors.append(f"第{i+1}个列定义必须是对象格式")
                        continue
                    
                    # 列的必需字段
                    if 'name' not in column:
                        errors.append(f"第{i+1}个列缺少 name 字段")
                    if 'type' not in column:
                        errors.append(f"第{i+1}个列缺少 type 字段")
                    
                    # 列名长度检查
                    if 'name' in column and len(column['name']) > 255:
                        errors.append(f"第{i+1}个列名过长（最大255字符）")
        
        if errors:
            table_key = f"{table_data.get('database', 'unknown')}.{table_data.get('name', 'unknown')}"
            self.validation_errors.extend([f"{table_key}: {error}" for error in errors])
            return False
        
        return True
    
    def parse_json_format(self, file_content: str) -> List[Dict[str, Any]]:
        """解析JSON格式的元数据"""
        try:
            data = json.loads(file_content)
            if 'tables' in data:
                return data['tables']
            elif isinstance(data, list):
                return data
            else:
                raise ValueError("JSON格式不正确，应包含 'tables' 数组或直接为表数组")
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON解析失败: {str(e)}")
    
    def parse_csv_format(self, file_content: str) -> List[Dict[str, Any]]:
        """解析CSV格式的元数据"""
        try:
            csv_reader = csv.DictReader(StringIO(file_content))
            tables = {}
            
            for row in csv_reader:
                db_name = row.get('database', '').strip()
                table_name = row.get('table_name', '').strip()
                column_name = row.get('column_name', '').strip()
                column_type = row.get('column_type', '').strip()
                column_comment = row.get('column_comment', '').strip()
                
                if not all([db_name, table_name, column_name, column_type]):
                    continue
                
                table_key = f"{db_name}.{table_name}"
                if table_key not in tables:
                    tables[table_key] = {
                        'database': db_name,
                        'name': table_name,
                        'columns': []
                    }
                
                tables[table_key]['columns'].append({
                    'name': column_name,
                    'type': column_type,
                    'comment': column_comment
                })
            
            return list(tables.values())
        except Exception as e:
            raise ValueError(f"CSV解析失败: {str(e)}")
    
    def parse_excel_format(self, file_content: bytes) -> List[Dict[str, Any]]:
        """解析Excel格式的元数据"""
        try:
            df = pd.read_excel(file_content)
            tables = {}
            
            for _, row in df.iterrows():
                db_name = str(row.get('database', '')).strip()
                table_name = str(row.get('table_name', '')).strip()
                column_name = str(row.get('column_name', '')).strip()
                column_type = str(row.get('column_type', '')).strip()
                column_comment = str(row.get('column_comment', '')).strip()
                
                if not all([db_name, table_name, column_name, column_type]):
                    continue
                
                table_key = f"{db_name}.{table_name}"
                if table_key not in tables:
                    tables[table_key] = {
                        'database': db_name,
                        'name': table_name,
                        'columns': []
                    }
                
                tables[table_key]['columns'].append({
                    'name': column_name,
                    'type': column_type,
                    'comment': column_comment
                })
            
            return list(tables.values())
        except Exception as e:
            raise ValueError(f"Excel解析失败: {str(e)}")
    
    def preview_import_data(self, file_content: Any, file_format: str) -> Dict[str, Any]:
        """预览导入数据"""
        self.validation_errors = []
        
        try:
            if file_format == 'json':
                tables = self.parse_json_format(file_content)
            elif file_format == 'csv':
                tables = self.parse_csv_format(file_content)
            elif file_format == 'excel':
                tables = self.parse_excel_format(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_format}")
            
            # 验证数据
            valid_tables = []
            for table in tables:
                if self.validate_table_data(table):
                    valid_tables.append(table)
            
            return {
                'total_tables': len(tables),
                'valid_tables': len(valid_tables),
                'invalid_tables': len(tables) - len(valid_tables),
                'validation_errors': self.validation_errors,
                'preview_data': valid_tables[:5],  # 只预览前5个表
                'sample_table': valid_tables[0] if valid_tables else None
            }
        
        except Exception as e:
            return {
                'total_tables': 0,
                'valid_tables': 0,
                'invalid_tables': 0,
                'validation_errors': [str(e)],
                'preview_data': [],
                'sample_table': None
            }
    
    def import_metadata(self, file_content: Any, file_format: str, import_mode: str = 'merge') -> Dict[str, Any]:
        """
        导入元数据
        
        Args:
            file_content: 文件内容
            file_format: 文件格式 (json, csv, excel)
            import_mode: 导入模式 (merge: 合并, replace: 替换)
        
        Returns:
            导入结果统计
        """
        self.validation_errors = []
        self.import_stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'errors': 0,
            'skipped': 0
        }
        
        try:
            # 解析数据
            if file_format == 'json':
                tables = self.parse_json_format(file_content)
            elif file_format == 'csv':
                tables = self.parse_csv_format(file_content)
            elif file_format == 'excel':
                tables = self.parse_excel_format(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_format}")
            
            self.import_stats['total'] = len(tables)
            
            # 使用事务确保数据一致性
            with transaction.atomic():
                for table_data in tables:
                    try:
                        if not self.validate_table_data(table_data):
                            self.import_stats['errors'] += 1
                            continue
                        
                        # 检查表是否已存在
                        existing_table = HiveTable.objects.filter(
                            database=table_data['database'],
                            name=table_data['name']
                        ).first()
                        
                        if existing_table:
                            if import_mode == 'merge':
                                # 合并模式：更新已存在的表
                                existing_table.columns = table_data['columns']
                                existing_table.save()
                                self.import_stats['updated'] += 1
                            else:
                                # 跳过已存在的表
                                self.import_stats['skipped'] += 1
                        else:
                            # 创建新表
                            HiveTable.objects.create(
                                database=table_data['database'],
                                name=table_data['name'],
                                columns_json=json.dumps(table_data['columns'])
                            )
                            self.import_stats['created'] += 1
                    
                    except Exception as e:
                        self.validation_errors.append(
                            f"{table_data.get('database', 'unknown')}.{table_data.get('name', 'unknown')}: {str(e)}"
                        )
                        self.import_stats['errors'] += 1
                        continue
            
            return {
                'success': True,
                'stats': self.import_stats,
                'errors': self.validation_errors
            }
        
        except Exception as e:
            return {
                'success': False,
                'stats': self.import_stats,
                'errors': [str(e)]
            }
    
    def get_import_template(self, format_type: str) -> Dict[str, Any]:
        """获取导入模板"""
        sample_data = {
            'tables': [
                {
                    'database': 'dwd_zbk',
                    'name': 'user_info',
                    'comment': '用户信息表',
                    'columns': [
                        {
                            'name': 'user_id',
                            'type': 'bigint',
                            'comment': '用户ID'
                        },
                        {
                            'name': 'user_name',
                            'type': 'string',
                            'comment': '用户姓名'
                        },
                        {
                            'name': 'create_time',
                            'type': 'timestamp',
                            'comment': '创建时间'
                        }
                    ]
                },
                {
                    'database': 'dwd_zbk', 
                    'name': 'order_info',
                    'comment': '订单信息表',
                    'columns': [
                        {
                            'name': 'order_id',
                            'type': 'bigint',
                            'comment': '订单ID'
                        },
                        {
                            'name': 'user_id',
                            'type': 'bigint',
                            'comment': '用户ID'
                        },
                        {
                            'name': 'amount',
                            'type': 'decimal(10,2)',
                            'comment': '订单金额'
                        }
                    ]
                }
            ]
        }
        
        if format_type == 'json':
            return {
                'content_type': 'application/json',
                'content': json.dumps(sample_data, ensure_ascii=False, indent=2)
            }
        elif format_type == 'csv':
            csv_content = "database,table_name,column_name,column_type,column_comment\n"
            # user_info 表数据
            csv_content += "dwd_zbk,user_info,user_id,bigint,用户ID\n"
            csv_content += "dwd_zbk,user_info,user_name,string,用户姓名\n"
            csv_content += "dwd_zbk,user_info,create_time,timestamp,创建时间\n"
            # order_info 表数据
            csv_content += "dwd_zbk,order_info,order_id,bigint,订单ID\n"
            csv_content += "dwd_zbk,order_info,user_id,bigint,用户ID\n"
            csv_content += "dwd_zbk,order_info,amount,decimal(10\\,2),订单金额\n"
            return {
                'content_type': 'text/csv',
                'content': csv_content
            }
        elif format_type == 'excel':
            # 生成Excel格式数据
            try:
                import io
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "元数据模板"
                
                # 添加表头
                headers = ['database', 'table_name', 'column_name', 'column_type', 'column_comment']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 添加示例数据
                row = 2
                for table in sample_data['tables']:
                    for column in table['columns']:
                        ws.cell(row=row, column=1, value=table['database'])
                        ws.cell(row=row, column=2, value=table['name'])
                        ws.cell(row=row, column=3, value=column['name'])
                        ws.cell(row=row, column=4, value=column['type'])
                        ws.cell(row=row, column=5, value=column['comment'])
                        row += 1
                
                # 保存到字节流
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_content = excel_buffer.getvalue()
                excel_buffer.close()
                
                return {
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'content': excel_content
                }
                
            except Exception as e:
                # 如果Excel生成失败，回退到JSON格式
                return {
                    'content_type': 'application/json',
                    'content': json.dumps(sample_data, ensure_ascii=False, indent=2)
                }
        else:
            return {
                'content_type': 'application/json',
                'content': json.dumps(sample_data, ensure_ascii=False, indent=2)
            }